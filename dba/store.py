"""Load the DBA session workbook into a queryable SQLite database.

Why SQLite instead of RAG: this sheet is 2,372 rows of *numeric telemetry*.
Questions like "which database burns the most CPU?" require summing every row —
retrieval can only return the handful of rows that look textually similar, which
produces confident wrong answers. So we compute with SQL and let the LLM only
translate the question and narrate the result.

The workbook is treated as a read-only snapshot. Nothing here ever writes back
to the source file, and :func:`run_query` refuses anything but SELECT.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path

# Columns that are numeric in meaning even though Excel stores them loosely.
_NUMERIC = {"CPU", "used_memory", "reads", "writes", "physical_reads",
            "session_id", "blocking_session_id"}

# Only these statements may ever run. An LLM writing SQL against a real
# database is a genuine risk, so the guard is a whitelist, not a blacklist.
_FORBIDDEN = re.compile(
    r"\b(insert|update|delete|drop|alter|create|replace|truncate|attach|"
    r"detach|pragma|vacuum|reindex)\b",
    re.IGNORECASE,
)

DEFAULT_XLSX = Path("data/dba/db_session.xlsx")
TABLE = "db_sessions"


class DBAError(RuntimeError):
    """Raised when the workbook can't be read or a query is rejected."""


@dataclass
class Column:
    name: str
    meaning: str = ""

    def describe(self) -> str:
        return f"{self.name}: {self.meaning}" if self.meaning else self.name


class SessionStore:
    """An in-memory SQLite mirror of the DBA workbook."""

    def __init__(self, xlsx: str | Path = DEFAULT_XLSX) -> None:
        self.xlsx = Path(xlsx)
        if not self.xlsx.exists():
            raise DBAError(f"Workbook not found: {self.xlsx}")
        self.conn = sqlite3.connect(":memory:", check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.columns: list[Column] = []
        self.row_count = 0
        self._load()

    # -- loading ------------------------------------------------------------
    def _load(self) -> None:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise DBAError(
                "openpyxl is required to read .xlsx files. "
                "Install it with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(self.xlsx, read_only=True, data_only=True)
        if TABLE not in wb.sheetnames:
            raise DBAError(f"Sheet {TABLE!r} not found in {self.xlsx.name}")

        rows = list(wb[TABLE].iter_rows(values_only=True))
        if not rows:
            raise DBAError("The session sheet is empty.")

        header = [str(h).strip() for h in rows[0] if h is not None]
        meanings = self._read_dictionary(wb)
        self.columns = [Column(h, meanings.get(h, "")) for h in header]

        cols_sql = ", ".join(
            f'"{h}" {"INTEGER" if h in _NUMERIC else "TEXT"}' for h in header
        )
        self.conn.execute(f"CREATE TABLE {TABLE} ({cols_sql})")

        payload = []
        for raw in rows[1:]:
            if raw is None or all(c is None for c in raw):
                continue
            record = []
            for name, cell in zip(header, raw):
                record.append(self._coerce(name, cell))
            payload.append(record)

        placeholders = ", ".join(["?"] * len(header))
        self.conn.executemany(
            f"INSERT INTO {TABLE} VALUES ({placeholders})", payload
        )
        self.conn.commit()
        self.row_count = len(payload)

    @staticmethod
    def _coerce(name: str, cell):
        """Excel cells arrive as mixed types; normalise for reliable SQL."""
        if cell is None:
            return None
        if name in _NUMERIC:
            try:
                return int(float(str(cell).strip()))
            except (TypeError, ValueError):
                return None
        return str(cell).strip()

    @staticmethod
    def _read_dictionary(wb) -> dict[str, str]:
        """The 'explanation' sheet documents each column — real gold for prompting."""
        out: dict[str, str] = {}
        if "explanation" not in wb.sheetnames:
            return out
        for row in wb["explanation"].iter_rows(values_only=True):
            if row and row[0]:
                key = str(row[0]).strip()
                val = " ".join(str(row[1] or "").split())
                out[key] = val
        return out

    # -- querying -----------------------------------------------------------
    def schema_prompt(self) -> str:
        """A compact schema + data dictionary for the text-to-SQL prompt."""
        lines = [
            f"Table: {TABLE}  ({self.row_count} rows)",
            "Columns (name: meaning):",
        ]
        for col in self.columns:
            lines.append(f"  - {col.describe()[:160]}")
        return "\n".join(lines)

    def run_query(self, sql: str, limit: int = 50) -> tuple[list[str], list[tuple]]:
        """Execute a single read-only SELECT and return (headers, rows)."""
        cleaned = sql.strip().rstrip(";").strip()
        if not cleaned:
            raise DBAError("Empty query.")
        if ";" in cleaned:
            raise DBAError("Only one statement may be run at a time.")
        if not re.match(r"^\s*(select|with)\b", cleaned, re.IGNORECASE):
            raise DBAError("Only SELECT queries are allowed.")
        # Strip string literals before scanning for dangerous keywords: a value
        # like WHERE name='update_job' is harmless, and blocking it would be a
        # confusing false positive.
        scannable = re.sub(r"'[^']*'", "''", cleaned)
        if _FORBIDDEN.search(scannable):
            raise DBAError("This query attempts to modify data, which is blocked.")

        try:
            cur = self.conn.execute(cleaned)
        except sqlite3.Error as exc:
            raise DBAError(f"SQL error: {exc}") from exc

        headers = [d[0] for d in cur.description or []]
        rows = [tuple(r) for r in cur.fetchmany(limit)]
        return headers, rows

    def sample(self, n: int = 3) -> tuple[list[str], list[tuple]]:
        return self.run_query(f"SELECT * FROM {TABLE} LIMIT {n}")


_CACHE: dict[str, SessionStore] = {}


def get_store(xlsx: str | Path = DEFAULT_XLSX) -> SessionStore:
    """Cached loader so the workbook is parsed once per process."""
    key = str(Path(xlsx).resolve())
    if key not in _CACHE:
        _CACHE[key] = SessionStore(xlsx)
    return _CACHE[key]
